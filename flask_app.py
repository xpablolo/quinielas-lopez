from flask import Flask, render_template
from openpyxl import load_workbook
from partidos_list import lista

app = Flask(__name__)

datos = "/home/pablolo/mysite/datos.xlsx"
resultados_xlsx = "/home/pablolo/mysite/Resultados.xlsx"
list_jornadas = []
book = load_workbook(datos)
sheet = book.active
for i in range(2, sheet.max_row+1):
    list_jornadas.append(sheet.cell(i,1).value)

@app.route("/")
def inicio():
    return render_template("inicio.html")


@app.route("/predicciones")
def predicciones():
    lista_jor = list_jornadas
    return render_template("predicciones.html", lista_jor=lista_jor)


@app.route("/resultados")
def resultados():
    book = load_workbook(resultados_xlsx)
    sheet = book.active
    return render_template("resultados.html", sheet=sheet)


@app.route("/predicciones/jornada-<int:week>")
def pred_jornada(week):
    book = load_workbook(datos)
    sheet = book.active
    index_jornada = list_jornadas.index(week)+1
    partidos_list = lista[index_jornada-1]
    book2 = load_workbook(resultados_xlsx)
    resultados = book2.active
    lista_jor = list_jornadas
    return render_template("jornada.html", sheet=sheet, partidos_list=partidos_list, jornada=week, index_jornada=index_jornada, resultados=resultados, lista_jor=lista_jor)


if __name__ == "__main__":
    app.run(debug=True)

